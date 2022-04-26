from openpyxl import load_workbook
from parse import *
from catalog.models import Size, Color, Collection, Carpet, DesignInColor, Design
from shipment.models import ShipmentItem

class PackingSheetProcessor:
    def __init__(self, shipment):
        """
        This class is used to handle the functionality of
        1. parsing the packing sheet in the shipment obj
        2. create the required model objects
        3. save the newly created model object on confirmation on the
            intermediary page.
        """
        self.shipment = shipment
        self.colors = Color.objects.all()
        self.sizes = Size.objects.all()
        self.designs = Design.objects.all()
        self.designColors = DesignInColor.objects.all()
        self.quality = Collection.objects.all()
        self.objects_to_be_created = []

    def confirm_creation(self):
        for obj in self.objects_to_be_created:
            obj.save()

    def parseColor(self, color):
        prs = parse("{} / {}", color)
        col, _ = self.colors.get_or_create(primary_color=prs[0], texture_color=prs[1])
        return col


    def parseSize(self, size):
        prs = parse("{:d} x {:d} {:w}", size, case_sensitive=False)
        if(prs is None):
            prs = parse("{:d}x{:d} {:w}", size, case_sensitive=False)

        siz, _ = self.sizes.get_or_create(width=prs[0]*1.0, length=prs[1]*1.0, shape=prs[2])
        return siz


    def process_shipment(self, qualityName, designName, color, size, ean, quantity):
        def _create():
            sizeInstance = self.parseSize(size)
            colorInstance = self.parseColor(color)
            quality, _ = self.quality.get_or_create(name=qualityName)
            design, _ = self.designs.get_or_create(name=designName, collection=quality)
            designInColor, _ = self.designColors.get_or_create(design=design, color=colorInstance)
            cpt, _ = Carpet.objects.get_or_create(designColor=designInColor, size=sizeInstance)
            shipmentItem, _ = ShipmentItem.objects.get_or_create(
                shipment=self.shipment, carpet=cpt, barcode=ean, quantity=quantity, )

        if(ean):
            if(ean.isnumeric()):
                _create()
                return 1
        return 0


    def process_packing_sheet(self):
        print("Shipment Created with packing sheet " + str(self.shipment.packing_sheet.url))
        workbook = load_workbook(self.shipment.packing_sheet)
        packing_sheet = workbook.active

        col_names = {}
        column_index = 0
        REQ_NAMES = set(['QUALITY', 'DESIGN', 'COLOR', 'SIZE', 'EAN CODE', 'PCS'])
        for column in packing_sheet.iter_cols(0, packing_sheet.max_column):
            print("processing top row")
            if column[0].value in REQ_NAMES:
                col_names[column[0].value] = column_index
            column_index += 1

        shipment_item_created_count = 0
        c = 1
        for row in packing_sheet.iter_rows(0, packing_sheet.max_row):
            print(f"processint row {c}")
            c += 1
            try:
                shipment_item_created_count += self.process_shipment(row[col_names['QUALITY']].value,
                                    row[col_names['DESIGN']].value,
                                    row[col_names['COLOR']].value,
                                    row[col_names['SIZE']].value,
                                    row[col_names['EAN CODE']].value,
                                    row[col_names['PCS']].value)
                print("in try")
            except Exception as e:
                print("exception found", e)
                return 0, e

        return shipment_item_created_count, None
