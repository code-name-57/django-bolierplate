from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def ReallyCreateShipmentItem(qualityName, designName, color, size, quantity, ean):
    print("Quality : "+qualityName+
        " Design name : "+designName+
        " Color : "+color+" Size : "+size+
        " Quantity :"+str(quantity))

def CreateShipmentItem(qualityName, designName, color, size, ean, quantity):
    if(qualityName and designName and color and size and quantity and ean ):
        ReallyCreateShipmentItem(qualityName, designName, color, size, quantity, ean)

def shipment_created(sender, instance, created, **kwargs):

    print("Shipment Created with packing sheet " + str(instance.packing_sheet.url))
    workbook = load_workbook('/home/avinash/code-name-57/django-bolierplate/media/shipments/Customer_Packing_List_with_EAN_Codes_ATDOplv.xlsx')
    packing_sheet = workbook.active

    col_names = {}
    column_index = 0 
    REQ_NAMES = set(['QUALTY', 'DESIGN', 'COLOR', 'SIZE', 'EAN CODE', 'PCS'])
    for column in packing_sheet.iter_cols(0, packing_sheet.max_column):
        if column[0].value in REQ_NAMES:
            col_names[column[0].value] = column_index
        column_index += 1



    from collections import defaultdict
    data = defaultdict(list)
    for row in packing_sheet.iter_rows(2, packing_sheet.max_row):
        for key, index in col_names.items():
            data[key].append(row[index].value)
            CreateShipmentItem(row[col_names['QUALTY']].value,
                            row[col_names['DESIGN']].value,
                            row[col_names['COLOR']].value,
                            row[col_names['SIZE']].value,
                            row[col_names['EAN CODE']].value,
                            row[col_names['PCS']].value)

    pass
